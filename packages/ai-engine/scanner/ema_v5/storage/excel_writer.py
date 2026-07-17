"""
EMA_V5 Excel Writer — Produces EMA_V5_SIGNALS.xlsx in isolated output directory.
Never touches existing Excel exports.
"""
from __future__ import annotations

import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Font = PatternFill = Alignment = Border = Side = None  # type: ignore
    logger.warning("openpyxl not installed — Excel export disabled")


# ── Column definitions — matches spec exactly ────────────────────
_COLUMNS = [
    ("Date", 12),
    ("Time", 10),
    ("Exchange", 10),
    ("Symbol", 14),
    ("Side", 8),
    ("Trend", 14),
    ("Current State", 16),
    ("EMA20", 12),
    ("EMA50", 12),
    ("EMA144", 12),
    ("EMA200", 12),
    ("Entry", 14),
    ("Stop Loss", 14),
    ("TP1", 14),
    ("TP2", 14),
    ("TP3", 14),
    ("Volume", 8),
    ("Confidence", 12),
    ("Reason", 24),
    ("Pattern", 18),
    ("Result", 8),
    ("PnL", 10),
    ("Hold Time", 10),
    ("Strategy Version", 14),
]

# ── Styles (only if openpyxl available) ──────────────────────────
if HAS_OPENPYXL:
    _HEADER_FONT = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
    _HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    _LONG_FONT = Font(name="Consolas", size=9, color="3FB950")
    _SHORT_FONT = Font(name="Consolas", size=9, color="F85149")
    _DEFAULT_FONT = Font(name="Consolas", size=9, color="C9D1D9")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="30363D"),
        right=Side(style="thin", color="30363D"),
        top=Side(style="thin", color="30363D"),
        bottom=Side(style="thin", color="30363D"),
    )
else:
    _HEADER_FONT = _HEADER_FILL = _LONG_FONT = _SHORT_FONT = _DEFAULT_FONT = _THIN_BORDER = None


class EMAv5ExcelWriter:
    """Produces EMA_V5_SIGNALS.xlsx — isolated output, never touches existing exports."""

    def __init__(self, output_dir: Optional[str] = None) -> None:
        _root = Path(output_dir) if output_dir else Path(__file__).resolve().parent.parent.parent.parent / "data"
        self._dir = _root / "ema_v5_exports"
        self._dir.mkdir(parents=True, exist_ok=True)
        self._filename = "EMA_V5_SIGNALS.xlsx"

    def _build_row(self, signal: Dict[str, Any]) -> List[Any]:
        """Map a signal dict to a flat row matching _COLUMNS."""
        return [
            signal.get("date", ""),
            signal.get("time", ""),
            signal.get("exchange", "Binance"),
            signal.get("symbol", ""),
            signal.get("side", ""),
            signal.get("trend", ""),
            signal.get("current_state", ""),
            round(signal.get("ema20", 0), 6),
            round(signal.get("ema50", 0), 6),
            round(signal.get("ema144", 0), 6),
            round(signal.get("ema200", 0), 6),
            round(signal.get("entry", 0), 6),
            round(signal.get("stop_loss", 0), 6),
            round(signal.get("tp1", 0), 6),
            round(signal.get("tp2", 0), 6),
            round(signal.get("tp3", 0), 6),
            "✅" if signal.get("volume") else "❌",
            f"{(signal.get('confidence', 0) or 0):.1f}%",
            signal.get("reason", ""),
            signal.get("pattern", ""),
            signal.get("result", "") or "—",
            f"{signal.get('pnl', 0):.4f}" if signal.get("pnl") else "—",
            f"{signal.get('hold_time', 0):.0f}m" if signal.get("hold_time") else "—",
            signal.get("strategy_version", "ema_v5"),
        ]

    def write(self, signals: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
        """Write all signals to Excel. Returns file path.
        
        Creates a new file each time (append mode via openpyxl).
        """
        if not HAS_OPENPYXL:
            logger.warning("Cannot write Excel — openpyxl not installed")
            return ""

        target = filename or self._filename
        filepath = self._dir / target

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EMA_V5_SIGNALS"

        # ── Header row ──
        for col_idx, (name, width) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # ── Data rows ──
        for row_idx, signal in enumerate(signals, 2):
            row_data = self._build_row(signal)
            side = signal.get("side", "")
            font = _LONG_FONT if side == "LONG" else _SHORT_FONT if side == "SHORT" else _DEFAULT_FONT

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = font
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

        # ── Freeze header ──
        ws.freeze_panes = "A2"

        # ── Auto-filter ──
        if signals:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(_COLUMNS))}{len(signals) + 1}"

        try:
            wb.save(str(filepath))
            logger.info("📊 EMA_V5 Excel written: {} ({} signals)", filepath, len(signals))
            return str(filepath)
        except Exception as e:
            logger.error("EMAv5 Excel write failed: {}", e)
            return ""

    def read(self, filename: Optional[str] = None) -> List[Dict[str, Any]]:
        """Read signals from existing Excel file."""
        if not HAS_OPENPYXL:
            return []

        target = filename or self._filename
        filepath = self._dir / target

        if not filepath.exists():
            return []

        try:
            wb = openpyxl.load_workbook(str(filepath), read_only=True)
            ws = wb.active

            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            signals = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                sig = dict(zip(headers, row))
                signals.append(sig)

            wb.close()
            return signals
        except Exception as e:
            logger.error("EMAv5 Excel read failed: {}", e)
            return []

    def get_path(self) -> str:
        """Return the full path to the Excel file."""
        return str(self._dir / self._filename)

    def exists(self) -> bool:
        """Check if Excel file exists."""
        return (self._dir / self._filename).exists()
